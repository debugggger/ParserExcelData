import os
import threading
import time

import pandas as pd

import openpyxl

from tkinter import *
from threading import Thread
from tkinter import filedialog
from pathlib import Path

from peoplesData import People

stopParsingThread = threading.Event()
checkFilesThread = threading.Event()


def startEvent():
    getChangeTime()
    global threadParsing
    threadParsing = Thread(target=parsingData)
    threadParsing.start()
    btnStop.grid(column=1, row=3)


def stopEvent():
    stopParsingThread.set()
    btnStop.grid_remove()

def browseFiles():
    global fileName
    fileName = filedialog.askopenfilename(initialdir="/",
                                          title="Select a File",
                                          filetypes=(("Excel files",
                                                      "*.xls*"),
                                                     ("all files",
                                                      "*.*")))
    getSheetName()
    filePath.insert(0, fileName)

def getSheetName():
    sheetName = ""
    if varGender.get() == 0:
        sheetName = 'Рез_Муж'
    if varGender.get() == 1:
        sheetName = 'Рез_Жен'
    return sheetName



def getReserveFile():
    try:
        reservePath = r'C:\Users' '\\' + os.getenv('USERNAME') + '\AppData\Roaming\Microsoft\Excel' '\\'
        name = (fileName.split('/')[-1]).split('.')[0]

        checkT = 0
        maxTime = 0.01
        truePath = ''
        for dir in os.listdir(reservePath):
            if len(name) <= len(dir):
                for i in range(len(name)):
                    if name[i] == dir[i]:
                        checkT = 1
                        truePath = reservePath + dir
                    else:
                        checkT = 0
                        break

            if checkT == 1:
                if os.path.getmtime(reservePath + dir) > maxTime:
                    maxTime = os.path.getmtime(reservePath + dir)
                    truePath = reservePath + dir + '\\'

        trueFile = ''
        maxTimeFile = 0.01
        for file in os.listdir(truePath):
            current = truePath + file
            if os.path.getmtime(current) > maxTimeFile:
                if current.split('.')[-1] == 'xlsb':
                    maxTimeFile = os.path.getmtime(current)
                    trueFile = current
                if current.split('.')[-1] == 'xls':
                    if len(current) >= len(trueFile):
                        maxTimeFile = os.path.getmtime(current)
                        trueFile = current


        eng = ''
        if trueFile.split('.')[-1] == 'xls':
            eng = 'xlrd'

        if trueFile.split('.')[-1] == 'xlsb':
            eng = 'pyxlsb'

        sheetName = getSheetName()
        df = pd.read_excel(trueFile, engine=eng, sheet_name=sheetName)
        nameTrueFile = (trueFile.split('/')[-1]).split('.')[0] + '.xlsx'
        df.to_excel(nameTrueFile)

        xlsx = openpyxl.load_workbook(nameTrueFile, data_only=True)
        sheetName = 'Sheet1'
        sheet = xlsx[sheetName]
        sheet.delete_cols(1, 1)
        xlsx.save(nameTrueFile)
        xlsx.close()
        return nameTrueFile
    except:
        return ''

def getChangeTime():
    global t1
    t1 = os.path.getmtime(fileName)

def parsingData():
    prevFile = ''
    while True:

        reserveFile = getReserveFile()

        if t1 != os.path.getmtime(fileName):
            if fileName.split('.')[-1] == 'xls':
                reservePath = r'C:\Users' '\\' + os.getenv('USERNAME') + '\AppData\Roaming\Microsoft\Excel' '\\'
                df = pd.read_excel(fileName, engine='xlrd', sheet_name=getSheetName())
                nameTrueFile = (fileName.split('/')[-1]).split('.')[0] + '.xlsx'
                patchXlsx = reservePath+nameTrueFile
                df.to_excel(patchXlsx)

                sheetName = 'Sheet1'
                xlsx = openpyxl.load_workbook(patchXlsx, data_only=True)
                sheet = xlsx.get_sheet_by_name(sheetName)
                sheet.delete_cols(1, 1)
                xlsx.save(patchXlsx)
                readData(patchXlsx, sheetName)
                os.remove(patchXlsx)
            else:
                readData(fileName, getSheetName())
            getChangeTime()

        if reserveFile != prevFile:
            prevFile = reserveFile
            readData(reserveFile, "Sheet1")
            #os.remove(reserveFile)

        if stopParsingThread.is_set():
            break
        time.sleep(10)

def readData(nameReadFile, sheetName):
    peoples = []

    extension = nameReadFile.split('.')[-1]
    if extension == 'xlsx':
        xlsx = openpyxl.load_workbook(nameReadFile, data_only=True)
        sheet = xlsx[sheetName]


        for cellObj in sheet[startPlace.get():lastPlace.get()]:
            currentPeople = People("", "", "", "", "", "", "", "", "", "", "")
            firstColumnInd = cellObj[0].column
            currentColumn = firstColumnInd
            for cell in cellObj:
                if currentColumn == firstColumnInd:
                    currentPeople.place = cell.value
                if currentColumn == firstColumnInd + 3:
                    currentPeople.name = cell.value
                if currentColumn == firstColumnInd + 4:
                    currentPeople.year = cell.value
                if currentColumn == firstColumnInd + 5:
                    currentPeople.discharge = cell.value
                if currentColumn == firstColumnInd + 6:
                    currentPeople.city = cell.value
                if currentColumn == firstColumnInd + 9:
                    currentPeople.school = cell.value
                if currentColumn == firstColumnInd + 10:
                    currentPeople.c1 = cell.value
                if currentColumn == firstColumnInd + 11:
                    currentPeople.c2 = cell.value
                if currentColumn == firstColumnInd + 12:
                    currentPeople.c3 = cell.value
                if currentColumn == firstColumnInd + 21:
                    currentPeople.seks = cell.value
                if currentColumn == firstColumnInd + 23:
                    currentPeople.total = cell.value
                currentColumn += 1
            if currentPeople.total is not None:
                peoples.append(currentPeople)

    for i in range(len(peoples)-1):
        for j in range(len(peoples)-i-1):
            cPeople = peoples[j]
            nPeople = peoples[j+1]
            if cPeople.getTotal() < nPeople.getTotal():
                peoples[j], peoples[j+1] = peoples[j+1], peoples[j]

    for i in peoples:
        People.showData(i)
    print('______________________________')


window = Tk()
window.title("")
width = window.winfo_screenwidth()
height = window.winfo_screenheight()
window.geometry('%dx%d' % (width/3, height/4))

lbl = Label(window, text="Путь к файлу")
lbl.grid(column=0, row=0)
filePath = Entry(window, width=10)
filePath.grid(column=1, row=0)
btn = Button(window, text="поиск", command=browseFiles)
btn.grid(column=2, row=0)

startPlaceLbl = Label(window, text="Первая ячейка таблицы")
startPlaceLbl.grid(column=0, row=1)
startPlace = Entry(window, width=10)
startPlace.grid(column=1, row=1)
lastPlaceLbl = Label(window, text="Последняя ячейка таблицы")
lastPlaceLbl.grid(column=2, row=1)
lastPlace = Entry(window, width=10)
lastPlace.grid(column=3, row=1)

varGender = IntVar()
varGender.set(0)
rbtnMan = Radiobutton(text='Мужские соревнования', variable=varGender, value=0)
rbtnWoman = Radiobutton(text='Женские соревнования', variable=varGender, value=1)
rbtnMan.grid(column=0, row=2)
rbtnWoman.grid(column=1, row=2)


btnStart = Button(window, text="начать", command=startEvent)
btnStart.grid(column=0, row=3)

btnStop = Button(window, text="остановить", command=stopEvent)
btnStop.grid_remove()


lblReserve = Label(window, text="Путь к папке")
lblReserve.grid(column=0, row=4)
filePathReserve = Entry(window, width=10)
filePathReserve.grid(column=1, row=4)
btnReserve = Button(window, text="поиск", command=browseFiles)
btnReserve.grid(column=2, row=4)



window.mainloop()