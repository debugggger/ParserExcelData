import os
import re
import string
import threading
import time
import pandas as pd
import openpyxl

from threading import Thread
from tkinter import filedialog

from bot import Bot
from peoplesData import People

from beautifultable import BeautifulTable

class Parser(object):
    def __init__(self, app, btnStop, filePath, filePathReserve, dataPlace, ManList, WomanList):
        self.xlsxFileMan = ''
        self.xlsxFileWoman = ''
        self.fileName = ''
        self.peoples = []
        self.prevExcDataM = []
        self.prevExcDataW = []
        self.cvalResMCol = 0
        self.cvalResMRow = 0
        self.finalResMCol = 0
        self.finalResMRow = 0
        self.prevReserveFile = ''
        self.sheet = ''
        self.tChOriFile = 0.0
        self.tChResFile = 0.0
        self.btnStop = btnStop
        self.filePath = filePath
        self.filePathReserve = filePathReserve
        self.lastPlace = dataPlace
        self.startPlace = 'A1'
        self.colInd = 0
        self.colLetter = ''
        self.manList = ManList
        self.womanList = WomanList
        self.app = app
        self.stopParsingThread = threading.Event()
        self.bot = Bot()

    def col2num(self, col):
        try:
            num = 0
            for c in col:
                if c in string.ascii_letters:
                    num = num * 26 + (ord(c.upper()) - ord('A')) + 1
            return num
        except:
            self.app.errorMessageAdr()

    def lastColLetter(self, col):
        try:
            letters = ''
            for c in col:
                if c in string.ascii_letters:
                    letters += c
            return letters
        except:
            self.app.errorMessageAdr()

    def startEvent(self):
        self.app.updAll()
        self.app.errorMsg.grid_remove()
        threadParsing = Thread(target=self.parsingData)
        threadParsing.start()
        self.btnStop.grid(column=1, row=6)
        self.colInd = self.col2num(self.lastPlace)
        self.colLetter = self.lastColLetter(self.lastPlace)


    def stopEvent(self):
        self.stopParsingThread.set()
        self.btnStop.grid_remove()

    def browseFiles(self):
        self.app.clearFileName()
        self.fileName = filedialog.askopenfilename(initialdir="/",
                                              title="Select a File",
                                              filetypes=(("Excel files",
                                                          "*.xls*"),
                                                         ("all files",
                                                          "*.*")))
        self.filePath.insert(0, self.fileName)
        if self.fileName != '':
            self.getChangeTime()
            reserveFile = self.getReserveFile()
            if reserveFile != '':
                if self.tChResFile > self.tChOriFile:
                    self.getPrevData(reserveFile, 'autoRecovery')
                if self.tChResFile <= self.tChOriFile:
                    self.getPrevData(self.fileName, 'handRecovery')
            else:
                self.getPrevData(self.fileName, 'handRecovery')

    def getReservePatch(self):
        reservePath = self.filePathReserve.get()
        name = (self.fileName.split('/')[-1]).split('.')[0]

        checkT = 0
        maxTime = 0.01
        truePath = ''
        for dir in os.listdir(reservePath):
            if len(name) <= len(dir):
                for i in range(len(name)):
                    if name[i] == dir[i]:
                        checkT = 1
                        truePath = reservePath + dir
                    else:
                        checkT = 0
                        break

            if checkT == 1:
                if os.path.getmtime(reservePath + dir) > maxTime:
                    maxTime = os.path.getmtime(reservePath + dir)
                    truePath = reservePath + dir + '\\'
        return truePath

    def getReserveFile(self):
        try:
            truePath = self.getReservePatch()

            trueFile = ''
            maxTimeFile = 0.01
            for file in os.listdir(truePath):
                current = truePath + file
                if os.path.getmtime(current) > maxTimeFile:
                    if current.split('.')[-1] == 'xlsb':
                        maxTimeFile = os.path.getmtime(current)
                        trueFile = current
                    if current.split('.')[-1] == 'xls':
                        if len(current) >= len(trueFile):
                            maxTimeFile = os.path.getmtime(current)
                            trueFile = current

            if self.prevReserveFile == trueFile:
                return ''
            return trueFile
        except:
            return ''


    def getXlsxReserveFile(self):
        trueFile = self.getReserveFile()
        if trueFile == '':
            return trueFile
        self.tChResFile = os.path.getmtime(trueFile)
        self.prevReserveFile = trueFile
        eng = ''
        if trueFile.split('.')[-1] == 'xls':
            eng = 'xlrd'

        if trueFile.split('.')[-1] == 'xlsb':
            eng = 'pyxlsb'

        if len(self.prevExcDataM) > 0:
            dfM = pd.read_excel(trueFile, engine=eng, sheet_name=self.manList)
            nameTrueFileM = (trueFile.split('/')[-1]).split('.')[0] + '.xlsx'
            dfM.to_excel(nameTrueFileM)
            xlsx = openpyxl.load_workbook(nameTrueFileM, data_only=True)
            sheetName = 'Sheet1'
            sheet = xlsx[sheetName]
            sheet.delete_cols(1, 1)
            index = 0
            for row in sheet[self.startPlace:self.lastPlace]:
                for cell in row:
                    if cell.value != self.prevExcDataM[index]:
                        if type(cell.value) == str:
                            if re.search(r'Unnamed', cell.value):
                                index += self.colInd
                                break
                        self.sheet = self.manList
                        sheet.title = self.sheet
                        xlsx.save(nameTrueFileM)
                        xlsx.close()
                        return nameTrueFileM
                    index += 1

        if len(self.prevExcDataW) > 0:
            dfW = pd.read_excel(trueFile, engine=eng, sheet_name=self.womanList)
            nameTrueFileW = (trueFile.split('/')[-1]).split('.')[0] + '.xlsx'
            dfW.to_excel(nameTrueFileW)

            xlsx = openpyxl.load_workbook(nameTrueFileW, data_only=True)
            sheetName = 'Sheet1'
            sheet = xlsx[sheetName]
            sheet.delete_cols(1, 1)
            index = 0
            for row in sheet[self.startPlace:self.lastPlace]:
                for cell in row:
                    if cell.value != self.prevExcDataW[index]:
                        if type(cell.value) == str:
                            if re.search(r'Unnamed', cell.value):
                                index += self.colInd
                                break
                        self.sheet = self.womanList
                        sheet.title = self.sheet
                        xlsx.save(nameTrueFileW)
                        xlsx.close()
                        return nameTrueFileW
                    index += 1
        return ''


    def getChangeTime(self):
        self.tChOriFile = os.path.getmtime(self.fileName)

    def parsingData(self):
        try:
            prevFile = ''
            while True:
                reserveFile = self.getXlsxReserveFile()
                isChange = 0

                if self.tChOriFile != os.path.getmtime(self.fileName):
                    if self.fileName.split('.')[-1] == 'xls':
                        if len(self.prevExcDataM) > 0:
                            sheetName = 'Sheet1'
                            if reserveFile == '':
                                sheetName = self.manList
                                reserveFile = self.xls2xlsx(self.fileName, sheetName)
                            xlsx = openpyxl.load_workbook(reserveFile, data_only=True)
                            sheet = xlsx[sheetName]
                            index = 0
                            isExit = 0
                            for row in sheet[self.startPlace:self.lastPlace]:
                                for cell in row:
                                    if cell.value != self.prevExcDataM[index]:
                                        if type(cell.value) == str:
                                            if re.search(r'Unnamed', cell.value):
                                                index += self.colInd
                                                break
                                        self.sheet = self.manList
                                        sheet.title = self.sheet
                                        xlsx.save(reserveFile)
                                        xlsx.close()
                                        self.readData(reserveFile, 'autoRecovery')
                                        isExit = 1
                                        isChange = 1
                                        break
                                    index += 1
                                if isExit == 1:
                                    break

                        if len(self.prevExcDataW) > 0 and isChange == 0:
                            sheetName = 'Sheet1'
                            if reserveFile == self.xlsxFileMan:
                                sheetName = self.womanList
                                reserveFile = self.xls2xlsx(self.fileName, sheetName)
                            xlsx = openpyxl.load_workbook(reserveFile, data_only=True)
                            sheet = xlsx[sheetName]
                            index = 0
                            isExit = 0
                            for row in sheet[self.startPlace:self.lastPlace]:
                                for cell in row:
                                    if cell.value != self.prevExcDataW[index]:
                                        if type(cell.value) == str:
                                            if re.search(r'Unnamed', cell.value):
                                                index += self.colInd
                                                break
                                        self.sheet = self.womanList
                                        sheet.title = self.sheet
                                        xlsx.save(reserveFile)
                                        xlsx.close()
                                        self.readData(reserveFile, 'autoRecovery')
                                        isExit = 1
                                        break
                                    index += 1
                                if isExit == 1:
                                    break
                    else:
                        self.readData(self.fileName, 'handRecovery')
                    self.getChangeTime()


                if reserveFile != prevFile and os.path.getmtime(self.fileName) < self.tChResFile and reserveFile != '':
                    prevFile = reserveFile
                    self.readData(reserveFile, 'autoRecovery')

                if self.stopParsingThread.is_set():
                    break
                time.sleep(10)
        except:
            self.app.errorMessageNoFile()

    def getMData(self, file):
        try:
            self.prevExcDataM = []
            xlsx = openpyxl.load_workbook(file, data_only=True)
            currentSheet = xlsx[self.manList]
            i = 1
            for row in currentSheet[self.startPlace:self.lastPlace]:
                j = 1
                for cell in row:
                    self.prevExcDataM.append(cell.value)
                    if type(cell.value) == str and re.search(r'????????', cell.value):
                        self.cvalResMCol = cell.column_letter
                        self.cvalResMRow = i + 2
                    if type(cell.value) == str and re.search(r'??????????', cell.value):
                        self.finalResMCol = cell.column_letter
                        self.finalResMRow = i + 2
                    j += 1
                i += 1
        except:
            self.app.errorMessageNoSheet()

    def getWData(self, file):
        try:
            self.prevExcDataW = []
            xlsx = openpyxl.load_workbook(file, data_only=True)
            currentSheet = xlsx[self.womanList]
            i = 1
            for row in currentSheet[self.startPlace:self.lastPlace]:
                j = 1
                for cell in row:
                    self.prevExcDataW.append(cell.value)
                    if type(cell.value) == str and re.search(r'????????', cell.value):
                        self.cvalResWCol = cell.column_letter
                        self.cvalResWRow = i + 2
                    if type(cell.value) == str and re.search(r'??????????', cell.value):
                        self.finalResWCol = cell.column_letter
                        self.finalResWRow = i + 2
                    j += 1
                i += 1
        except:
            self.app.errorMessageNoSheet()

    def xls2xlsx(self, file, currentSheet):
        df = pd.read_excel(file, engine='xlrd', sheet_name=currentSheet)
        self.patchXlsx = self.getReservePatch() + currentSheet+'-'+(file.split('/')[-1]).split('.')[0] + '.xlsx'
        df.to_excel(self.patchXlsx)
        sheetName = 'Sheet1'
        xlsx = openpyxl.load_workbook(self.patchXlsx, data_only=True)
        sheet = xlsx.get_sheet_by_name(sheetName)
        sheet.delete_cols(1, 1)
        sheet.title = currentSheet
        xlsx.save(self.patchXlsx)
        return self.patchXlsx

    def getPrevData(self, file, mode):
        tempFile = file
        if mode == 'handRecovery':
            if file.split('.')[-1] == 'xls':
                self.xlsxFileMan = self.xls2xlsx(file, self.manList)
                tempFile = self.xlsxFileMan
            self.getMData(tempFile)
            if file.split('.')[-1] == 'xls':
                self.xlsxFileWoman = self.xls2xlsx(file, self.womanList)
                tempFile = self.xlsxFileWoman
            self.getWData(tempFile)
        if mode == 'autoRecovery':
            xl = pd.ExcelFile(file)
            if xl.sheet_names[0] == self.manList:
                if file.split('.')[-1] == 'xls':
                    self.xlsxFileMan = self.xls2xlsx(file, self.manList)
                    tempFile = self.xlsxFileMan
                self.getMData(tempFile)
            if xl.sheet_names[0] == self.womanList:
                if file.split('.')[-1] == 'xls':
                    self.xlsxFileWoman = self.xls2xlsx(file, self.womanList)
                    tempFile = self.xlsxFileWoman
                self.getWData(tempFile)

    def readData(self, nameReadFile, mode):
        global thisPeople
        self.peoples = []
        xlsx = openpyxl.load_workbook(nameReadFile, data_only=True)
        isChange = 0
        index = 0
        isExit = 0
        i = 1
        changedRow = 0
        startRow = 0
        stage = ''
        startPlace = self.startPlace
        lastPlace = self.lastPlace

        if mode == 'autoRecovery' and self.sheet != '':
            currentSheet = xlsx[self.sheet]

            for row in currentSheet[self.startPlace:self.lastPlace]:
                for cell in row:
                    if self.sheet == self.manList and len(self.prevExcDataM) > 0:
                        if cell.value != self.prevExcDataM[index]:
                            if type(cell.value) == str:
                                if re.search(r'Unnamed', cell.value):
                                    index += self.colInd
                                    break
                            isChange = 1
                            self.sheet = self.manList
                            if i >= self.cvalResMRow:
                                startPlace = str(self.cvalResMCol) + str(self.cvalResMRow)
                                lastPlace = self.lastPlace
                                isExit = 1
                                startRow = self.cvalResMRow
                                changedRow = i
                                stage = '????????????????????????, ??????????????'
                                break
                            if i >= self.finalResMRow and i < self.cvalResMRow:
                                startPlace = str(self.finalResMCol) + str(self.finalResMRow)
                                lastPlace = self.colLetter + str(self.cvalResMRow - 3)
                                isExit = 1
                                startRow = self.finalResMRow
                                changedRow = i
                                stage = '??????????, ??????????????'
                                break
                    if self.sheet == self.womanList and len(self.prevExcDataW) > 0:
                        if cell.value != self.prevExcDataW[index]:
                            if type(cell.value) == str:
                                if re.search(r'Unnamed', cell.value):
                                    index += self.colInd
                                    break
                            isChange = 1
                            self.sheet = self.womanList
                            if i >= self.cvalResWRow:
                                startPlace = str(self.cvalResWCol) + str(self.cvalResWRow)
                                lastPlace = self.lastPlace
                                isExit = 1
                                startRow = self.cvalResWRow
                                changedRow = i
                                stage = '????????????????????????, ??????????????'
                                break
                            if i >= self.finalResWRow and i < self.cvalResWRow:
                                startPlace = str(self.finalResWCol) + str(self.finalResWRow)
                                lastPlace = self.colLetter + str(self.cvalResWRow - 3)
                                isExit = 1
                                startRow = self.finalResWRow
                                changedRow = i
                                stage = '??????????, ??????????????'
                                break
                    index += 1
                i += 1
                if isExit == 1:
                    break

        if mode == 'handRecovery':
            if len(self.prevExcDataM) > 0:
                currentSheet = xlsx[self.manList]
                for row in currentSheet[self.startPlace:self.lastPlace]:
                    for cell in row:
                        if cell.value != self.prevExcDataM[index]:
                            isChange = 1
                            self.sheet = self.manList
                            if i >= self.cvalResMRow:
                                startPlace = str(self.cvalResMCol) + str(self.cvalResMRow)
                                lastPlace = self.lastPlace
                                isExit = 1
                                startRow = self.cvalResMRow
                                changedRow = i
                                stage = '????????????????????????, ??????????????'
                                break
                            if i >= self.finalResMRow and i < self.cvalResMRow:
                                startPlace = str(self.finalResMCol) + str(self.finalResMRow)
                                lastPlace = self.colLetter + str(self.cvalResMRow - 3)
                                isExit = 1
                                startRow = self.finalResMRow
                                changedRow = i
                                stage = '??????????, ??????????????'
                                break
                        index += 1
                    i += 1
                    if isExit == 1:
                        break

            if isChange == 0 and len(self.prevExcDataW) > 0:
                i = 1
                currentSheet = xlsx[self.womanList]
                index = 0
                for row in currentSheet[self.startPlace:self.lastPlace]:
                    for cell in row:
                        if cell.value != self.prevExcDataW[index]:
                            isChange = 1
                            self.sheet = self.womanList
                            if i >= self.cvalResWRow:
                                startPlace = str(self.cvalResWCol) + str(self.cvalResWRow)
                                lastPlace = self.lastPlace
                                isExit = 1
                                startRow = self.cvalResWRow
                                changedRow = i
                                stage = '????????????????????????, ??????????????'
                                break
                            if i >= self.finalResWRow and i < self.cvalResWRow:
                                startPlace = str(self.finalResWCol) + str(self.finalResWRow)
                                lastPlace = self.colLetter + str(self.cvalResWRow - 3)
                                isExit = 1
                                startRow = self.finalResWRow
                                changedRow = i
                                stage = '??????????, ??????????????'
                                break
                        index += 1
                    i += 1
                    if isExit == 1:
                        break


        if isChange == 1:
            sheet = xlsx[self.sheet]
            currentRow = 0
            thisPeople = People("", "", "", "", "", "", "", "", "", "", "", "", "")
            for cellObj in sheet[startPlace:lastPlace]:
                currentRow += 1
                currentPeople = People("", "", "", "", "", "", "", "", "", "", "", "", "")
                firstColumnInd = cellObj[0].column
                currentColumn = firstColumnInd
                for cell in cellObj:
                    if currentColumn == firstColumnInd + 1:
                        currentPeople.place = cell.value
                    if currentColumn == firstColumnInd + 3:
                        currentPeople.name = cell.value
                    if currentColumn == firstColumnInd + 4:
                        currentPeople.year = cell.value
                    if currentColumn == firstColumnInd + 5:
                        currentPeople.discharge = cell.value
                    if currentColumn == firstColumnInd + 6:
                        currentPeople.city = cell.value
                    if currentColumn == firstColumnInd + 9:
                        currentPeople.school = cell.value
                    if currentColumn == firstColumnInd + 10:
                        currentPeople.c1 = cell.value
                    if currentColumn == firstColumnInd + 11:
                        currentPeople.c2 = cell.value
                    if currentColumn == firstColumnInd + 12:
                        currentPeople.c3 = cell.value
                    if currentColumn == firstColumnInd + 17:
                        currentPeople.turns1 = cell.value
                    if currentColumn == firstColumnInd + 20:
                        currentPeople.turns2 = cell.value
                    if currentColumn == firstColumnInd + 22:
                        currentPeople.secBalls = cell.value
                    if currentColumn == firstColumnInd + 23:
                        currentPeople.total = cell.value
                    currentColumn += 1
                if currentRow == changedRow - startRow + 1:
                    thisPeople = currentPeople
                if currentPeople.total is not None:
                    self.peoples.append(currentPeople)

            for i in range(len(self.peoples) - 1):
                for j in range(len(self.peoples) - i - 1):
                    cPeople = self.peoples[j]
                    nPeople = self.peoples[j + 1]
                    if cPeople.getTotal() < nPeople.getTotal():
                        self.peoples[j], self.peoples[j + 1] = self.peoples[j + 1], self.peoples[j]

            startNumbers = []
            names = []
            totals = []
            startNumbers.append("??????????")
            names.append("??????????????")
            totals.append("??????????????????")
            print(stage)
            message = stage + "\n"
            print('______________________________')
            for i in self.peoples:
                if People.getTotal(i) != 0:
                    # print(People.getPlace(i), " ", People.getName(i), " ", People.getYear(i), " ", People.getDischarge(i),
                    #       " ", People.getCity(i), " ", People.getSchool(i), " ", People.getC1(i), " ", People.getC2(i),
                    #       " ", People.getC3(i), " ",
                    #       People.getTurns1(i), People.getTurns2(i), People.getSecBalls(i), " ", People.getTotal(i))

                    startNumbers.append(str(People.getPlace(i)))
                    names.append(str(People.getName(i)))
                    totals.append(str(People.getTotal(i)))

            startNumbers2 = []
            startNumbersLength = 0
            for i in startNumbers:
                if len(i) > startNumbersLength:
                    startNumbersLength = len(i)
            for i in range(len(startNumbers)):
                fixLength = 0
                if len(startNumbers[i]) < startNumbersLength:
                    fixLength = (startNumbersLength - len(startNumbers[i]))*2+1
                c = 0
                while c < fixLength:
                    c += 1
                    startNumbers[i] += " "
                startNumbers2.append(startNumbers[i])

            names2 = []
            namesLength = 0
            for i in names:
                if len(i) > namesLength:
                    namesLength = len(i)
            for i in range(len(names)):
                fixLength = 0
                if len(names[i]) < namesLength:
                    fixLength = (namesLength - len(names[i])) * 2 + 1
                c = 0
                while c < fixLength:
                    c += 1
                    names[i] += " "
                names2.append(names[i])

            totals2 = []
            totalsLength = 0
            for i in totals:
                if len(i) > totalsLength:
                    totalsLength = len(i)
            for i in range(len(names)):
                fixLength = 0
                if len(totals[i]) < totalsLength:
                    fixLength = (totalsLength - len(totals[i])) * 2 + 1
                c = 0
                while c < fixLength:
                    c += 1
                    totals[i] += " "
                totals2.append(totals[i])

            myTable = ""
            for i in range(len(startNumbers)):
                myTable += startNumbers2[i] + '|' + names2[i] + '|' + totals2[i] + "\n"

            print(myTable)

            print('______________________________')

            message += "\n?????????????????? ?????????? - " + str(thisPeople.getPlace()) + "\n?????????????? - " \
                       + str(thisPeople.getName()) + "\n?????? ???????????????? - " + str(thisPeople.getYear())\
                        + "\n???????????? - " +  str(thisPeople.getDischarge()) + "\n?????????? - " +  str(thisPeople.getCity())\
                        + "\n?????????? - " +  str(thisPeople.getSchool() + "\n?????????? ???? ????????????????" + "\n?????????? 1 - " + str(thisPeople.getC1())\
                        + "\n?????????? 2 - " + str(thisPeople.getC2()) + "\n?????????? 3 - " +  str(thisPeople.getC3())\
                        + "\n?????????? ???? ????????????" + "\n?????????? 1 - " + str(thisPeople.getTurns1()) + "\n?????????? 2 - " + str(thisPeople.getTurns2())\
                        + "\n?????????? ???? ?????????? - " + str(thisPeople.getSecBalls()) + "\n?????????? - " + str(thisPeople.getTotal()))

            self.bot.send_telegram(message)

            self.bot.send_telegram(myTable)

            self.getPrevData(nameReadFile, mode)
